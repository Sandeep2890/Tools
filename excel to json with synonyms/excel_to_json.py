# Program to convert excel to json format with 3 fileds brand, colour and information about the product
# By receiving input during the Program excecution.
import openpyxl as op
import json


def make_json(excel_path, json_path):
    wb_obj = op.load_workbook(excel_path.strip())
    print("Worksheets in the workbook are :" + str(wb_obj.sheetnames))
    sheet1 = wb_obj["productdetails"]
    max_col = sheet1.max_column
    max_row = sheet1.max_row
    sheet2 = wb_obj['synonyms']

    # master list to store dictionary data of each excel row
    data = []
    myjson = {"customerid": cid, "requesttype": rtype}

    for i in range(2, max_row + 1):
        mydict = {}
        fielddata = []
        synonyms = []          # list to store synonyms
        for k in range(2, sheet2.max_column):          # To store synonym to synonym list
            synonym = sheet2.cell(i, k).value
            if synonym is not None:
                synonyms.append(synonym)
        for j in range(1, max_col + 1):                # To store extract data from excel sheet
            key = sheet1.cell(row=1, column=j).value
            value = sheet1.cell(row=i, column=j).value
            # if cell is empty
            if not value:
                # copy value from row 2
                value = sheet1.cell(row=2, column=j).value
                # if value in row2 is also empty
            if not value:
                value = "value not assigned"

            # creating the json relevant data structure
            if key == "name":
                mydict["name"] = value
                mydict["description"] = "some description coming from backend(python) side about this"
                mydict["synonyms"]=synonyms

            fields = ["brand", "color", "info"]
            if key in fields:
                fieldnamepair = {"name": key, "value": value}
                fielddata.append(fieldnamepair)
                mydict["fields"] = fielddata

            if key == "category":
                mydict["category"] = value

            if key == "subcategory":
                mydict["subcategory"] = value

        data.append(mydict)
    catalogue = [{"lang": language, "products": data}]
    with open(json_path.strip(), "w") as fout:
        myjson["catalogue"] = catalogue
        json.dump(myjson, fout, indent=3)
    return "Successfully built"


__name__ == "__main__"
excel_path = input("Enter the path for excel file with product details which is in xlsx format: ")
json_path = input("Enter the path for json file with extension .json: ")
cid = input("Enter customer id: ")
rtype = input("Enter request type: ")
language = input("Enter language: ")
res = make_json(excel_path, json_path)
print(res)
