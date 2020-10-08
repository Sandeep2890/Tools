import argparse
import requests
import json
import openpyxl as op

# 'http://52.47.198.128:8085/v1/ldnlu/ecom/customer/smobiles/job/103'
data = []
parser = argparse.ArgumentParser()
parser.add_argument('file', type=str, help='Configure file name')
args = parser.parse_args()
f = args.file
fp=open(f, 'r')
for line in fp:
    if line !="\n":
        if "# " not in line:
            data.append(line.strip())
lang=data[0]
url=data[1]
excel_path= data[5]   # "C:\\Users\\Sandeep vittal pai\\OneDrive\\Desktop\\linkedin_lifedata.xlsx"
utt_file=data[2]   # "D:\\Growth Falcons Documents\\utterances.txt"
out_file=data[4]
expected_file= data[3]  # "C:\\Users\\Sandeep vittal pai\\OneDrive\\Desktop\\utterances.txt"
wb_obj = op.load_workbook(excel_path)
s=wb_obj.active
s.title="MTT_output"
s.cell(1, 1).value="utterance"
s.cell(1, 2).value="output"
s.cell(1, 3).value="Actual output"
s.cell(1, 4).value="comparision result"
fout = open(out_file, "w")
fout.truncate()
fout.close()
utterances=[]
f=open(utt_file, 'r')
for line in f:
    if line !="\n":
        utterances.append(line.strip())
if len(utterances)==0:
    print("Text file "+'"'+utt_file+'"'+" is empty")
    exit()
data.clear()
f = open(expected_file, "r")
for line in f:
    if line !="\n":
        data.append(line.strip())
for i in range(0, len(utterances)):
    final_url = url+"?utterence="+utterances[i].strip('\n')+"&lang="+lang
    rg=requests.get(final_url)
    response = rg.json()
    value ={"fields": response["fields"]}
    s.cell(row=i + 2, column=1).value = utterances[i]
    s.cell(row=i + 2, column=2).value= json.dumps(value).strip("\n")
    s.cell(row=i + 2, column=3).value = data[i]
    if data[i].replace(" ","")==json.dumps(value).strip("\n").replace(" ",""):
        s.cell(i+2, 4).value="Pass"
    else:
        s.cell(i + 2, 4).value = "Fail"
wb_obj.save(excel_path)
if i==len(utterances)-1:
    print("Test competed data is stored in excel path provided")
else:
    print("incomplete testing process")
