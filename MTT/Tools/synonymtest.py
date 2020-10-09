import argparse
import requests
import json
import openpyxl as op
from openpyxl.styles import Color, PatternFill

parser = argparse.ArgumentParser()
parser.add_argument('file', type=str, help='Configure file name with input data')
args = parser.parse_args()
f = args.file
input_data = {}                # to store input data
fp = open(f, 'r')
for line in fp:
    if line != "\n":
        if "# " not in line:
            line=line.strip("\n")
            key=line.split("=")
            input_data[key[0].strip()]=key[1].strip()

wb_obj = op.load_workbook(input_data["excel_path"])
s = wb_obj.active
s.title = "MTT_output"
s.cell(1, 1).value = "utterance"
s.cell(1, 2).value = "output"
s.cell(1, 3).value = "Actual output"
s.cell(1, 4).value = "comparison result"

fout = open(input_data["out_file"], "w")
fout.truncate()
fout.close()

f = open(input_data["utt_file"], 'r')
utterences = []                      # To store the utterences
for line in f:
    if line != "\n":
        utterences.append(line.strip())
if len(utterences) == 0:
    print("Text file "+'"'+input_data["utt_file"]+'"'+" is empty")
    exit()

data1=[]          # to store the output of get request
data2=[]          # to store the expected output of get request
for i in range(0, len(utterences)):
    final_url = input_data["url"] + "?utterence=" + utterences[i].strip('\n')+"&lang="+input_data["lang"]
    rg = requests.get(final_url)
    response = rg.json()
    if input_data.get("out_file") is not None:
        out_json = json.dumps(response, indent=2)
        with open(input_data["out_file"], 'a') as fout:
            fout.write(final_url+'\n')
            fout.write(out_json+'\n'+'\n'+'\n')

    value ={"fields": response.get("fields","not found")}
    s.cell(row=i + 2, column=1).value = utterences[i]
    s.cell(row=i + 2, column=2).value = json.dumps(value).strip("\n")
    data1.append(json.dumps(value).strip("\n"))
fout.close()
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
if input_data.get("expected_file") is not None:
    f = open(input_data["expected_file"], "r")
    for line in f:
        if line != "\n":
            data2.append(line.strip())
    for i in range(0, len(utterences)):
        value = s.cell(i + 2, 2)
        s.cell(row=i + 2, column=3).value = data2[i]
        if data1[i].replace(" ", "") == data2[i].replace(" ",""):
            s.cell(i+2, 4).value = "pass"
            s.cell(i + 2, 5).fill = green_fill
        else:
            s.cell(i + 2, 4).value = "Fail"
            s.cell(i + 2, 5).fill = red_fill
wb_obj.save(input_data["excel_path"])

if i == len(utterences)-1:
    print("Test competed data is stored in excel path provided")
else:
    print("incomplete testing process")